"""
Excel Styles, Constants, and Formatting Helpers
================================================
All openpyxl styling objects and utility functions used by the report generator.
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================
# SEVERITY STYLES
# ============================================================
SEVERITY_STYLES = {
    "CRITICAL": {"fill": "FF0000", "font": "FFFFFF", "bold": True},
    "HIGH":     {"fill": "FF6600", "font": "FFFFFF", "bold": True},
    "MEDIUM":   {"fill": "FFCC00", "font": "000000", "bold": True},
    "LOW":      {"fill": "92D050", "font": "000000", "bold": False},
    "INFO":     {"fill": "BDD7EE", "font": "000000", "bold": False},
}


# ============================================================
# FONT OBJECTS
# ============================================================
HEADER_FONT = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
TITLE_FONT = Font(name='Calibri', bold=True, size=18, color='1F4E79')
SUBTITLE_FONT = Font(name='Calibri', bold=True, size=13, color='2E75B6')
SECTION_FONT = Font(name='Calibri', bold=True, size=11, color='1F4E79')
BODY_FONT = Font(name='Calibri', size=10, color='333333')
BODY_FONT_BOLD = Font(name='Calibri', bold=True, size=10, color='333333')
LINK_FONT = Font(name='Calibri', size=10, color='0563C1', underline='single')
NUM_FONT = Font(name='Calibri', size=10, color='666666')
CODE_FONT = Font(name='Consolas', size=9, color='333333')
SUMMARY_LABEL_FONT = Font(name='Calibri', bold=True, size=11, color='44546A')
SUMMARY_VALUE_FONT = Font(name='Calibri', size=11, color='333333')


# ============================================================
# FILL OBJECTS
# ============================================================
HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
ZEBRA_FILL_1 = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
ZEBRA_FILL_2 = PatternFill(start_color='F2F7FB', end_color='F2F7FB', fill_type='solid')
SECTION_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')


# ============================================================
# ALIGNMENT OBJECTS
# ============================================================
WRAP_ALIGN = Alignment(wrap_text=True, vertical='top')
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
CENTER_TOP = Alignment(horizontal='center', vertical='top')
LEFT_TOP = Alignment(horizontal='left', vertical='top', wrap_text=True)
RIGHT_TOP = Alignment(horizontal='right', vertical='top')


# ============================================================
# BORDER OBJECTS
# ============================================================
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
)
HEADER_BORDER = Border(
    left=Side(style='thin', color='16365C'), right=Side(style='thin', color='16365C'),
    top=Side(style='thin', color='16365C'), bottom=Side(style='medium', color='16365C')
)


# ============================================================
# UTILITY FUNCTIONS
# ============================================================
def severity_fill(sev):
    s = SEVERITY_STYLES.get(sev, SEVERITY_STYLES["INFO"])
    return PatternFill(start_color=s["fill"], end_color=s["fill"], fill_type='solid')


def severity_font(sev):
    s = SEVERITY_STYLES.get(sev, SEVERITY_STYLES["INFO"])
    return Font(name='Calibri', bold=s["bold"], size=10, color=s["font"])


def style_header_row(ws, col_count, row=1):
    ws.row_dimensions[row].height = 28
    for col in range(1, col_count + 1):
        c = ws.cell(row=row, column=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER_ALIGN
        c.border = HEADER_BORDER
    ws.freeze_panes = f'A{row + 1}'
    ws.auto_filter.ref = f"A{row}:{get_column_letter(col_count)}{row}"


def apply_zebra_and_borders(ws, max_row, max_col, data_start=2):
    """Apply zebra striping, borders, alignment, and font to data rows."""
    for r in range(data_start, max_row + 1):
        zebra = ZEBRA_FILL_1 if (r - data_start) % 2 == 0 else ZEBRA_FILL_2
        ws.row_dimensions[r].height = 22
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            cell.font = BODY_FONT
            if not cell.fill or cell.fill.start_color.index in (None, '00000000', 0):
                cell.fill = zebra
            if c == 1:
                cell.alignment = CENTER_TOP
                cell.font = NUM_FONT
            elif c == 4:
                cell.alignment = CENTER_TOP
            else:
                cell.alignment = LEFT_TOP


def apply_borders(ws, max_row, max_col):
    """Legacy wrapper — uses zebra striping."""
    apply_zebra_and_borders(ws, max_row, max_col)


def color_severity_col(ws, col_idx, max_row):
    for r in range(2, max_row + 1):
        cell = ws.cell(row=r, column=col_idx)
        val = str(cell.value).upper() if cell.value else ''
        for sev in SEVERITY_STYLES:
            if sev in val:
                cell.fill = severity_fill(sev)
                cell.font = severity_font(sev)
                cell.alignment = CENTER_TOP
                break


def color_priority_col(ws, col_idx, max_row):
    for r in range(2, max_row + 1):
        cell = ws.cell(row=r, column=col_idx)
        val = str(cell.value) if cell.value else ''
        cell.alignment = CENTER_TOP
        if 'P0' in val:
            cell.fill = severity_fill("CRITICAL"); cell.font = severity_font("CRITICAL")
        elif 'P1' in val:
            cell.fill = severity_fill("HIGH"); cell.font = severity_font("HIGH")
        elif 'P2' in val:
            cell.fill = severity_fill("MEDIUM"); cell.font = severity_font("MEDIUM")
        elif 'P3' in val or 'P4' in val:
            cell.fill = severity_fill("LOW"); cell.font = severity_font("LOW")

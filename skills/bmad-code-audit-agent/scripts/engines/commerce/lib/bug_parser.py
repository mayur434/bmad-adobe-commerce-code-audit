"""
Bug Report Excel Parser
=========================
Parses bug reports from Excel (.xlsx) files into the structured dict
used by BRDAnalysisEngine for bug impact analysis.

Expected Excel columns (case-insensitive, order-independent):
  Bug ID, Title, Severity, Description, Steps to Reproduce,
  Expected Behavior, Actual Behavior, Reported By, Environment,
  Error Logs, Suspected Modules, Suspected Files, Suspected Functions
"""

import os
import re

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


def parse_bug_excel(filepath):
    """Parse a bug report Excel file and return structured dict.

    Returns dict with key: bugs (list of bug dicts)
    """
    if not filepath or not os.path.isfile(filepath):
        print(f"   ⚠️  Bug report file not found: {filepath}")
        return {"bugs": []}

    if load_workbook is None:
        print("   ❌ openpyxl is required to parse Excel bug reports")
        return {"bugs": []}

    ext = os.path.splitext(filepath)[1].lower()
    if ext not in ('.xlsx', '.xlsm'):
        print(f"   ⚠️  Unsupported bug report format: {ext}. Expected .xlsx")
        return {"bugs": []}

    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        print(f"   ❌ Failed to open bug report: {e}")
        return {"bugs": []}

    bugs = []
    ws = wb.active

    # Read header row and build column index
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        wb.close()
        return {"bugs": []}

    col_map = _build_column_map(header_row)

    # Parse data rows
    for row in ws.iter_rows(min_row=2, values_only=True):
        bug = _parse_bug_row(row, col_map)
        if bug:
            bugs.append(bug)

    wb.close()
    return {"bugs": bugs}


def _build_column_map(header_row):
    """Build a mapping from normalized column name to column index."""
    col_map = {}
    aliases = {
        "bug_id": ["bug id", "bug_id", "id", "ticket", "ticket id", "issue id", "jira"],
        "title": ["title", "summary", "bug title", "issue title", "subject"],
        "severity": ["severity", "priority", "sev", "level"],
        "description": ["description", "details", "bug description", "issue description"],
        "steps": ["steps to reproduce", "steps", "repro steps", "reproduction steps", "reproduce"],
        "expected": ["expected behavior", "expected", "expected result", "expected outcome"],
        "actual": ["actual behavior", "actual", "actual result", "actual outcome"],
        "reported_by": ["reported by", "reporter", "reported_by", "author", "created by"],
        "environment": ["environment", "env", "found in", "detected in"],
        "error_logs": ["error logs", "error_logs", "logs", "error", "stack trace", "exception"],
        "modules": ["suspected modules", "modules", "suspected_modules", "affected modules", "module"],
        "files": ["suspected files", "files", "suspected_files", "affected files", "file path", "file"],
        "functions": ["suspected functions", "functions", "suspected_functions", "method", "methods", "function"],
        "status": ["status", "state", "bug status"],
    }

    for idx, header in enumerate(header_row):
        if header is None:
            continue
        normalized = str(header).strip().lower()
        for key, names in aliases.items():
            if normalized in names:
                col_map[key] = idx
                break

    return col_map


def _cell_str(row, col_map, key):
    """Get cell value as string, handling None."""
    idx = col_map.get(key)
    if idx is None or idx >= len(row):
        return ""
    val = row[idx]
    return str(val).strip() if val is not None else ""


def _split_csv(value):
    """Split comma/newline separated values."""
    if not value:
        return []
    items = re.split(r'[,\n]+', value)
    return [item.strip() for item in items if item.strip()]


def _parse_bug_row(row, col_map):
    """Parse a single row into a bug dict."""
    bug_id = _cell_str(row, col_map, "bug_id")
    title = _cell_str(row, col_map, "title")

    # Skip empty rows
    if not bug_id and not title:
        return None

    severity = _cell_str(row, col_map, "severity").lower()
    # Normalize severity values
    sev_map = {
        "p1": "critical", "blocker": "critical", "1": "critical",
        "p2": "high", "major": "high", "2": "high",
        "p3": "medium", "normal": "medium", "moderate": "medium", "3": "medium",
        "p4": "low", "minor": "low", "trivial": "low", "4": "low",
    }
    severity = sev_map.get(severity, severity) if severity else "medium"

    description = _cell_str(row, col_map, "description")
    steps_raw = _cell_str(row, col_map, "steps")
    expected = _cell_str(row, col_map, "expected")
    actual = _cell_str(row, col_map, "actual")
    reported_by = _cell_str(row, col_map, "reported_by")
    environment = _cell_str(row, col_map, "environment")
    error_logs = _cell_str(row, col_map, "error_logs")

    # Parse steps (numbered lines or newline-separated)
    steps = []
    if steps_raw:
        for line in steps_raw.split('\n'):
            line = line.strip()
            cleaned = re.sub(r'^\d+[\.\)]\s*', '', line)
            if cleaned:
                steps.append(cleaned)

    # Parse suspected areas
    modules = _split_csv(_cell_str(row, col_map, "modules"))
    files = _split_csv(_cell_str(row, col_map, "files"))
    functions = _split_csv(_cell_str(row, col_map, "functions"))

    return {
        "id": bug_id or f"BUG-{hash(title) % 10000:04d}",
        "title": title,
        "description": description,
        "steps_to_reproduce": steps,
        "expected_behavior": expected,
        "actual_behavior": actual,
        "severity": severity,
        "reported_by": reported_by,
        "environment": environment,
        "error_logs": error_logs,
        "suspected_area": {
            "modules": modules,
            "files": files,
            "functions": functions,
        },
    }
